"""
Build professional Excel workbook: Unified Philosophical Framework
MGA, moral systems thesis, Triadic Hardness, smooth-discrete, GCD-Projection, Inevitability of Arithmetic.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

TITLE_FONT = Font(name='Cambria', size=18, bold=True, color='1B2A4A')
SUBTITLE_FONT = Font(name='Cambria', size=13, italic=True, color='4A5568')
SECTION_FONT = Font(name='Cambria', size=14, bold=True, color='2D3748')
THEOREM_FONT = Font(name='Cambria', size=12, bold=True, color='1A365D')
BODY_FONT = Font(name='Cambria', size=11, color='2D3748')
HEADER_FONT = Font(name='Cambria', size=11, bold=True, color='FFFFFF')
QED_FONT = Font(name='Cambria', size=11, bold=True, color='276749')
REMARK_FONT = Font(name='Cambria', size=11, italic=True, color='718096')
RESULT_FONT = Font(name='Cambria', size=12, bold=True, color='276749')
NOVEL_FONT = Font(name='Cambria', size=11, bold=True, color='9B2C2C')

HEADER_FILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
THEOREM_FILL = PatternFill(start_color='EBF8FF', end_color='EBF8FF', fill_type='solid')
PROOF_FILL = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
SUCCESS_FILL = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')
NOVEL_FILL = PatternFill(start_color='FED7E2', end_color='FED7E2', fill_type='solid')
WARN_FILL = PatternFill(start_color='FEFCBF', end_color='FEFCBF', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E0'), right=Side(style='thin', color='CBD5E0'),
    top=Side(style='thin', color='CBD5E0'), bottom=Side(style='thin', color='CBD5E0'))
BOTTOM_BORDER = Border(bottom=Side(style='medium', color='1B2A4A'))
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
ME = 'H'

def scw(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

def wt(ws, r, t):
    ws.merge_cells(f'A{r}:{ME}{r}')
    c = ws[f'A{r}']; c.value = t; c.font = TITLE_FONT; c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 35; return r+1

def wst(ws, r, t):
    ws.merge_cells(f'A{r}:{ME}{r}')
    c = ws[f'A{r}']; c.value = t; c.font = SUBTITLE_FONT; c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 22; return r+1

def wsec(ws, r, t):
    ws.merge_cells(f'A{r}:{ME}{r}')
    c = ws[f'A{r}']; c.value = t; c.font = SECTION_FONT; c.alignment = Alignment(vertical='center'); c.border = BOTTOM_BORDER
    ws.row_dimensions[r].height = 28; return r+1

def wthm(ws, r, l, s):
    ws.merge_cells(f'A{r}:{ME}{r}')
    c = ws[f'A{r}']; c.value = f'{l}. {s}'; c.font = THEOREM_FONT; c.fill = THEOREM_FILL; c.alignment = WRAP
    ws.row_dimensions[r].height = max(30, 15*(len(s)//80+1)); return r+1

def wl(ws, r, t, font=None, fill=None):
    ws.merge_cells(f'A{r}:{ME}{r}')
    c = ws[f'A{r}']; c.value = t; c.font = font or BODY_FONT; c.fill = fill or PROOF_FILL; c.alignment = WRAP
    ws.row_dimensions[r].height = max(18, 15*(len(str(t))//90+1)); return r+1

def wq(ws, r): return wl(ws, r, '∎', font=QED_FONT)
def wr(ws, r, t): return wl(ws, r, t, font=REMARK_FONT)
def wres(ws, r, t): return wl(ws, r, t, font=RESULT_FONT, fill=SUCCESS_FILL)
def wb_(ws, r): ws.row_dimensions[r].height = 8; return r+1

def th(ws, r, headers, col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=col+i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER_WRAP; c.border = THIN_BORDER
    ws.row_dimensions[r].height = 25; return r+1

def tr(ws, r, vals, col=1, alt=False, fills=None):
    for i, v in enumerate(vals):
        c = ws.cell(row=r, column=col+i, value=v)
        c.font = BODY_FONT; c.alignment = CENTER_WRAP; c.border = THIN_BORDER
        if fills and i < len(fills) and fills[i]: c.fill = fills[i]
        elif alt: c.fill = ALT_ROW_FILL
    return r+1

wb = openpyxl.Workbook()
W = [5, 15, 15, 15, 15, 15, 15, 15]

# ──────────────────────────────────────────────
# SHEET 1: MORAL SYSTEMS THESIS
# ──────────────────────────────────────────────
ws = wb.active; ws.title = 'I. Moral Systems Thesis'; ws.sheet_properties.tabColor = '553C9A'; scw(ws, W)

r = 3
r = wt(ws, r, 'INTRINSIC LOSS IN CLAIMED-PREDICTIVE MORAL SYSTEMS')
r = wb_(ws, r)
r = wst(ws, r, 'A Formal Information-Theoretic Analysis')
r += 2

r = wsec(ws, r, 'CORE THESIS')
r = wthm(ws, r, 'Central Claim',
    'Systems claiming predictive moral derivation while operating reactively produce intrinsic loss through two mechanisms: (1) Signal attenuation — real feedback delayed or obscured. (2) Noise injection — false causality attributed to phenomena with no moral information.')
r = wb_(ws, r)

r = wsec(ws, r, 'INFORMATION-THEORETIC FORMULATION')
r = wl(ws, r, 'Signal attenuation = H(M_a | M_c): uncertainty about the ACTUAL mechanism given the CLAIM.')
r = wl(ws, r, 'Noise injection = H(M_c | M_a): information in the CLAIM not present in REALITY.')
r = wl(ws, r, 'These are the two terms of Variation of Information: VI(M_c, M_a) = H(M_a|M_c) + H(M_c|M_a).')
r = wres(ws, r, 'The thesis\'s two loss mechanisms ARE the variation of information. The thesis was an implicit information-theoretic framework.')
r = wb_(ws, r)

r = wsec(ws, r, 'FALSIFIABILITY TESTS')
r = wl(ws, r, '1. Orthogonal metrics: Do positions correlate with irrelevant variables?')
r = wl(ws, r, '2. Catastrophic pressure: Same institution, different response times based on visibility.')
r = wl(ws, r, '3. Consequence-correlated position adjustment demonstrates reactive operation.')
r = wb_(ws, r)

r = wsec(ws, r, 'SCOPE RESTRICTION IS A CHANNEL PROPERTY')
r = wthm(ws, r, 'Key Insight',
    'The thesis addresses only "expressed and contested" positions. In MGA terms: the public discourse channel is optimized for I(T; Y_social) — information relevant to social coordination — not I(T; Y_moral). Universal prohibitions have zero mutual information with the social coordination variable.')
r = wl(ws, r, 'This resolves the anticipated criticism "you only look at contested positions" — yes, because that\'s what the channel transmits.')
r = wb_(ws, r)

r = wsec(ws, r, 'CONANT-ASHBY REFRAME')
r = wthm(ws, r, 'Reframe',
    'These systems are GOOD REGULATORS of the WRONG VARIABLE. Not broken, not corrupt — competently managing social coordination while claiming to manage moral derivation. The gap is structural, not personal.')
r = wb_(ws, r)

r = wsec(ws, r, 'CASE STUDIES')
r = th(ws, r, ['Case', 'Claimed Mechanism', 'Actual Mechanism', 'Loss Type'], col=2)
cases = [
    ('Usury prohibition reversal', 'Moral law from scripture', 'Economic necessity adaptation', 'Signal attenuation'),
    ('Salem witch trials', 'Divine revelation of evil', 'Social panic + property disputes', 'Noise injection'),
    ('Early capitalism/laissez-faire', 'Natural law (predictive)', 'Reactive market coordination', 'Both'),
    ('Prohibition (alcohol)', 'Moral imperative', 'Social coordination + temperance movement', 'Signal attenuation'),
]
for i, v in enumerate(cases):
    r = tr(ws, r, list(v), col=2, alt=(i%2==1))


# ──────────────────────────────────────────────
# SHEET 2: MGA FORMALIZATION
# ──────────────────────────────────────────────
ws2 = wb.create_sheet('II. MGA Formalization'); ws2.sheet_properties.tabColor = '2D3748'; scw(ws2, W)

r = 2
r = wt(ws2, r, 'MECHANISM GAP ANALYSIS (MGA)')
r = wst(ws2, r, 'Self-Sealing Theorem & Gap Ecology')
r += 1

r = wsec(ws2, r, 'SELF-SEALING THEOREM')
r = wthm(ws2, r, 'Theorem',
    'Systems that control their own mechanism measurement exhibit monotonically non-decreasing gaps, as a consequence of the data processing inequality.')
r = wl(ws2, r, 'Model: A system with actual mechanism M_a (drifts) and claimed mechanism M_c (updated through self-controlled channel).')
r = wl(ws2, r, 'Self-sealing condition: the channel is intentionally lossy (β > 0).')
r = wl(ws2, r, 'Measure: VI(M_c, M_a) = variation of information between claim and reality.')
r = wb_(ws2, r)

r = wsec(ws2, r, 'SIMULATION RESULTS')
r = wthm(ws2, r, 'Key Findings',
    'v1: VI over sliding windows noisy. v2: Agreement rate decreases with β. v3 (Non-stationary): Growing complexity OR degrading channel recovers monotonic growth prediction.')
r = th(ws2, r, ['Mode', 'Description', 'Result', 'Prediction Confirmed?'], col=2)
sim_data = [
    ('Stationary', 'Fixed state space, fixed channel', 'Steady-state gap (not growing)', 'Partial'),
    ('Growing states', 'State space expands over time', 'Gap grows monotonically', '✅ Yes'),
    ('Degrading channel', 'Observation quality decreases', 'Gap grows monotonically', '✅ Yes'),
    ('Both (ACE-CL-007)', 'Complexity + decay together', 'Strongest growth', '✅ Yes'),
]
for i, v in enumerate(sim_data):
    fills = [None]*4
    if '✅' in v[3]: fills[3] = SUCCESS_FILL
    r = tr(ws2, r, list(v), col=2, alt=(i%2==1), fills=fills)

r += 2
r = wsec(ws2, r, 'GAP ECOLOGY')
r = wthm(ws2, r, 'Structural Attractor',
    'The moral-authority gap doesn\'t exist in isolation. It is embedded in a symbiotic cluster with educational, legal, and political authority gaps.')
r = wl(ws2, r, 'Above an interaction-strength threshold, this cluster has a structural attractor that temporary reform cannot escape.')
r = wl(ws2, r, 'Single-institution reform FAILS because the cluster re-establishes equilibrium.')
r = wl(ws2, r, 'Historical pattern: moral reform happens across institutions simultaneously (Reformation, Enlightenment, civil rights).')
r = wl(ws2, r, 'Explanation: You need to disrupt the cluster as a whole.')
r = wb_(ws2, r)

r = wsec(ws2, r, 'GAP CLUSTER')
r = th(ws2, r, ['Institution', 'Claimed Mechanism', 'Actual Mechanism'], col=2)
gap_cluster = [
    ('Moral authority', 'Derives moral truth', 'Coordinates social behavior'),
    ('Educational authority', 'Teaches critical thinking', 'Transmits orthodoxy'),
    ('Legal authority', 'Objective justice', 'Maintains social order'),
    ('Political authority', 'Represents constituents', 'Maintains power structures'),
]
for i, v in enumerate(gap_cluster):
    r = tr(ws2, r, list(v), col=2, alt=(i%2==1))


# ──────────────────────────────────────────────
# SHEET 3: TRIADIC HARDNESS & SMOOTH-DISCRETE
# ──────────────────────────────────────────────
ws3 = wb.create_sheet('III. Complexity Theory'); ws3.sheet_properties.tabColor = '276749'; scw(ws3, W)

r = 2
r = wt(ws3, r, 'TRIADIC HARDNESS PRINCIPLE & SMOOTH-DISCRETE DUALITY')
r += 1

r = wsec(ws3, r, 'TRIADIC HARDNESS PRINCIPLE')
r = wthm(ws3, r, 'Conjecture 2.1',
    '(|π⁻¹(k)| > poly(|k|)) ∧ (SR(S) > log|k|) ∧ (I_loss > log|k|) ⟹ P is NP-hard or physically constrained.')
r = wl(ws3, r, 'Translation: Problems with high underdetermination, super-logarithmic self-reference, and significant information loss are computationally hard.')
r = wb_(ws3, r)

r = wsec(ws3, r, 'VERIFICATION ON d(n)')
r = th(ws3, r, ['Property', 'Condition', 'Status', 'Evidence'], col=2)
triadic = [
    ('Underdetermination', '|F_k| > poly', '✅ Verified', '|F₁₂₅₅| = 69, fibers unbounded'),
    ('Self-Reference', 'SR(d) > log|k|', '✅ Verified', 'SR(d) = 2 (φ,τ require factorization)'),
    ('Information Loss', 'I_loss > log|k|', '✅ Verified', 'Up to 10.264 bits measured'),
]
for i, v in enumerate(triadic):
    fills = [None, None, SUCCESS_FILL, None]
    r = tr(ws3, r, list(v), col=2, alt=(i%2==1), fills=fills)

r += 2
r = wsec(ws3, r, 'THREE-DOMAIN CONVERGENCE')
r = th(ws3, r, ['Property', 'SAT (Comp. Sci.)', 'QM (Physics)', 'Diophantine (Math)'], col=2)
convergence = [
    ('Underdetermination', 'SAT solution space', 'Wave function superposition', 'Diophantine solution space'),
    ('Self-Reference', 'Variable cycles in clauses', 'Measurement basis choice', 'Recursive structure'),
    ('Information Loss', 'Choosing assignment', 'Quantum collapse', 'Gap formation'),
    ('Constraint', 'NP-hardness', 'Physical measurement limit', 'Impossibility'),
]
for i, v in enumerate(convergence):
    r = tr(ws3, r, list(v), col=2, alt=(i%2==1))

r += 2
r = wsec(ws3, r, 'SMOOTH-DISCRETE DUALITY')
r = wthm(ws3, r, 'Theorem 12.5',
    'SMOOTH-FIBER-INVERSION ∈ P: d_smooth(x) = x − x·e^(−γ)/ln(x) − ln(x) is invertible via binary search in O(log k + log(1/ε)) time.')
r = wl(ws3, r, 'The smooth version: continuous everywhere, strictly increasing for x > e^γ ≈ 1.78, invertible.')
r = wl(ws3, r, 'The discrete version: only defined at integers, has gaps, non-invertible (multiple preimages), no polynomial algorithm known.')
r = wres(ws3, r, 'Discretization creates computational complexity. The smooth version is in P; the discrete version may be NP-hard.')
r = wb_(ws3, r)

r = wsec(ws3, r, 'LAWVERE\'S UNIFICATION')
r = wthm(ws3, r, 'Theorem (Lawvere 1969)',
    'In a Cartesian closed category, if ∃ point-surjective e: A → (A → B), then every f: B → B has a fixed point. Contrapositive: if some f has no fixed point, then no such surjection exists.')
r = wl(ws3, r, 'Instances: Cantor (ℝ uncountable), Gödel (incompleteness), Turing (halting problem), Yanofsky (complexity theory).')
r = wl(ws3, r, 'Our extension: Self-referential underdetermined systems with info loss instantiate Lawvere in the category of computational problems.')


# ──────────────────────────────────────────────
# SHEET 4: GCD-PROJECTION & INEVITABILITY
# ──────────────────────────────────────────────
ws4 = wb.create_sheet('IV. GCD-Projection & Arithmetic'); ws4.sheet_properties.tabColor = '9B2C2C'; scw(ws4, W)

r = 2
r = wt(ws4, r, 'GCD-PROJECTION CONVERGENCE THEOREM')
r += 1

r = wsec(ws4, r, 'BRIDGE IDENTITY')
r = wthm(ws4, r, 'Theorem',
    'cos²(θ*) = 1/φ where θ* = arccos(1/√φ) ≈ 38.17°. The worst-case Euclidean algorithm contraction rate equals the alternating projection contraction rate at this angle.')
r = wl(ws4, r, 'Connects Lamé\'s theorem (1844) on Euclidean algorithm to von Neumann\'s alternating projection convergence (1949).')
r = wl(ws4, r, 'Through a previously unnamed angle involving the golden ratio φ.')
r = wb_(ws4, r)

r = wsec(ws4, r, 'PROVEN RESULTS')
r = wthm(ws4, r, 'Arithmetic Friedrichs Angle',
    'θ_A(a,b) = arccos(b^{−1/(2E(a,b))}) assigns a geometric angle to integer pairs.')
r = wthm(ws4, r, 'Fibonacci Supremum',
    'sup θ_A = θ*, achieved asymptotically by consecutive Fibonacci pairs. E/D ratio converges from 1.196 to 1.041.')
r = wthm(ws4, r, 'Modular Reduction Homomorphism',
    'gcd(ab, n) = gcd(gcd(a,n)·gcd(b,n), n) for ALL positive a,b,n. Proof: p-adic valuation (4 cases). Verified: 66,344 tests, 0 failures.')
r = wthm(ws4, r, 'Filter Composition',
    'C_p ∘ C_q = C_{lcm(p,q)}, redundancy = gcd(p,q). Proven in DFT domain.')

r += 2
r = wt(ws4, r, 'THE INEVITABILITY OF ARITHMETIC')
r = wst(ws4, r, 'Structural Attractor Analysis of Ontology Engine Results')
r += 1

r = wsec(ws4, r, 'THE QUESTION')
r = wl(ws4, r, 'The ontology engine starts with ~75 logical primitives and independently derives all 7 Peano axioms through 5 distinct pathways:')
r = th(ws4, r, ['#', 'Pathway', 'Starting Primitives', 'Route to Arithmetic'], col=2)
pathways = [
    ('1', 'Unary counting', 'σ, ∅', 'Direct successor construction'),
    ('2', 'Self-reference', 'Self-application', 'Recursive enumeration'),
    ('3', 'Construction', 'Type builders', 'Structural induction'),
    ('4', 'Modal logic', '□, ◊', 'Necessity of successor'),
    ('5', 'Modal-negational', '⇒, ✗, ⊖', 'Implication + negation + subtraction → addition'),
]
for i, v in enumerate(pathways):
    r = tr(ws4, r, list(v), col=2, alt=(i%2==1))

r += 1
r = wsec(ws4, r, 'THE STRONG CLAIM: ARITHMETIC IS A STRUCTURAL ATTRACTOR')
r = wthm(ws4, r, 'Thesis',
    'If 5 independent logical pathways converge on the same algebraic structure, the simplest explanation is that the structure is an ATTRACTOR in the space of formal systems. Any sufficiently expressive logical system will encounter arithmetic.')
r = wl(ws4, r, 'This is structural realism: mathematical structures exist as attractors, not as objects in a separate realm (Platonism) or arbitrary conventions (formalism).')
r = wb_(ws4, r)

r = wsec(ws4, r, 'IMPLICATIONS')
r = wl(ws4, r, '1. Mathematical truth is neither discovered nor invented — it is INEVITABLE.')
r = wl(ws4, r, '2. The foundations of mathematics are OVERDETERMINED: theorems are reachable from MANY starting axioms.')
r = wl(ws4, r, '3. Foundations are scaffolding, not load-bearing. The structure holds itself up.')
r = wl(ws4, r, '4. Wigner\'s "unreasonable effectiveness" dissolves: physics and math both select for structural stability.')
r = wb_(ws4, r)

r = wsec(ws4, r, 'OPEN QUESTION')
r = wthm(ws4, r, 'Question',
    'Is grammar a structural attractor the way arithmetic is? If any sufficiently complex information-processing system converges on grammar-like patterns, then language is not a human invention but a structural inevitability.')
r = wl(ws4, r, 'The Intel Neural Substrate (5K neurons, no backprop) generates grammatically correct English from structural dynamics alone.')
r = wl(ws4, r, 'If grammar IS an attractor: language is structurally inevitable. If not: language is contingent in a way arithmetic isn\'t.')


# ──────────────────────────────────────────────
# SHEET 5: CROSS-PROJECT SYNTHESIS
# ──────────────────────────────────────────────
ws5 = wb.create_sheet('V. Cross-Project Synthesis'); ws5.sheet_properties.tabColor = '744210'; scw(ws5, W)

r = 2
r = wt(ws5, r, 'CROSS-PROJECT SYNTHESIS')
r = wst(ws5, r, 'Connections between all mathematical and philosophical work')
r += 1

r = wsec(ws5, r, 'THE COMPLETE RESEARCH MAP')
r = th(ws5, r, ['Project', 'Domain', 'Core Result', 'Status'], col=2)
proj_map = [
    ('Deficiency Function', 'Number Theory', '20+ theorems on d(n) = n − φ(n) − τ(n)', 'Paper-ready'),
    ('Dynamical Horizon', 'Physics/Cosmology', 'Wave DM confirmed: 81% cores (p<10⁻²⁵)', 'Paper-ready'),
    ('Ontology Engine', 'Math/AI', '5 novel findings, 0 false positives, ρ irreducible', 'Active'),
    ('Moral Systems', 'Philosophy', 'Intrinsic loss in claimed-predictive systems', 'Multiple versions'),
    ('MGA', 'Info Theory', 'Self-sealing theorem + gap ecology', 'Simulated'),
    ('GCD-Projection', 'Number Theory', 'cos²(θ*)=1/φ, connects Lamé to von Neumann', 'Publishable'),
    ('Inevitability of Arithmetic', 'Phil. of Math', '5 pathways → arithmetic is structural attractor', 'Active'),
    ('Intel Neural Substrate', 'Computational Neuro', '5K neurons, no backprop: arithmetic + language', 'Working'),
    ('Neuro-Symbolic LLM', 'AI', '305K+ triples, 11 expression types, formal reasoning', 'Working'),
    ('Triadic Hardness', 'Complexity Theory', 'Underdetermination + SR + Info Loss → NP-hard', '85-90% confidence'),
    ('Smooth-Discrete Duality', 'Complexity Theory', 'Discretization creates computational complexity', 'Validated (8/8 tests)'),
]
for i, v in enumerate(proj_map):
    r = tr(ws5, r, list(v), col=2, alt=(i%2==1))

r += 2
r = wsec(ws5, r, 'DEEP CONNECTIONS')
r = wthm(ws5, r, 'Connection 1: Dynamical Horizon ↔ Smooth-Discrete',
    'Both identify a threshold where continuous structure becomes effectively discrete. DH: Planck scale. Complexity: discretization of smooth functions. Same phenomenon in physics and computation.')
r = wthm(ws5, r, 'Connection 2: Ontology Engine ↔ Inevitability of Arithmetic',
    'The engine proves computationally what the philosophical analysis argues: arithmetic is structurally inevitable. Five convergent pathways from ~75 primitives.')
r = wthm(ws5, r, 'Connection 3: MGA ↔ Moral Systems Thesis',
    'MGA formalizes the thesis\'s core move. Signal attenuation + noise injection = VI(M_c, M_a). The thesis was an implicit information-theoretic framework.')
r = wthm(ws5, r, 'Connection 4: Deficiency Function ↔ Triadic Hardness',
    'd(n) serves as concrete example: underdetermined (fibers unbounded), self-referential (SR=2), lossy (log₂|F_k| bits). Smooth version in P, discrete may be NP-hard.')
r = wthm(ws5, r, 'Connection 5: ρ Irreducibility ↔ Relational Structuralism',
    'The engine discovers that relations (ρ) are the sole irreducible primitive. Mathematics is about relations, not objects or truth. This is structural realism empirically demonstrated.')
r = wthm(ws5, r, 'Connection 6: Intel Substrate ↔ Grammar as Attractor',
    '5K self-organizing neurons generate grammatical English without backprop. If grammar is a structural attractor like arithmetic, the substrate is evidence.')
r = wb_(ws5, r)

r = wsec(ws5, r, 'THE UNIFYING THEME')
r = wres(ws5, r, 'Across all projects: STRUCTURE EMERGES FROM CONSTRAINTS. Whether the constraints are physical (Planck scale), logical (ontology primitives), informational (self-sealing channels), or computational (discretization) — the result is the same: stable structures that any sufficiently complex system converges toward.')


# ──────────────────────────────────────────────
out = r'D:\claude\proofs\Philosophical_Framework.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Sheets: {wb.sheetnames}')
