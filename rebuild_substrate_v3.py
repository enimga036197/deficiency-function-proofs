"""
Substrate V3: Bigger, clearer, no merged cells blocking input.
100 neurons. Clean layout. Input cells that WORK.
Everything labeled. Nothing cramped.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
import random
random.seed(77)

# Styles
TF = Font(name='Cambria', size=14, bold=True, color='1B2A4A')
SF = Font(name='Cambria', size=11, bold=True, color='2D3748')
HF = Font(name='Cambria', size=8, bold=True, color='FFFFFF')
BF = Font(name='Cambria', size=9, color='2D3748')
FF = Font(name='Consolas', size=9, color='1A365D')
RF = Font(name='Cambria', size=10, bold=True, color='276749')
HFILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FEFCBF', end_color='FEFCBF', fill_type='solid')
ACT_FILL = PatternFill(start_color='BEE3F8', end_color='BEE3F8', fill_type='solid')
EXC_FILL = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')
INH_FILL = PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid')
OUT_FILL = PatternFill(start_color='E9D8FD', end_color='E9D8FD', fill_type='solid')
TB = Border(left=Side('thin','E2E8F0'), right=Side('thin','E2E8F0'),
            top=Side('thin','E2E8F0'), bottom=Side('thin','E2E8F0'))
C = Alignment(horizontal='center', vertical='center')

NN = 100  # 100 neurons
NI = 16   # 16 input
NH = 64   # 64 hidden
NO = 20   # 20 output

print(f"Building V3 substrate: {NN} neurons ({NI}in, {NH}hid, {NO}out)")

wb = openpyxl.load_workbook(r'D:\claude\proofs\Deficiency_Research_Instrument.xlsx')
if 'Neural Substrate' in wb.sheetnames:
    del wb['Neural Substrate']

ws = wb.create_sheet('Neural Substrate', 8)
ws.sheet_properties.tabColor = '9B2C2C'

def nlabel(j):
    if j < NI: return f'I{j}'
    elif j < NI+NH: return f'H{j-NI}'
    else: return f'O{j-NI-NH}'

def ncol(j): return get_column_letter(j + 2)

# ═══ LAYOUT — no merged cells, clear sections ═══
# Col A = labels, B onwards = neurons
ws.column_dimensions['A'].width = 14
for j in range(NN):
    ws.column_dimensions[ncol(j)].width = 4.2

# Row 1: Title (in col A only, no merge)
ws['A1'].value = f'NEURAL SUBSTRATE — {NN} neurons live'
ws['A1'].font = TF

# Row 2: Architecture info
ws['A2'].value = f'{NI}in {NH}hid {NO}out | Sigmoid | F9=tick'
ws['A2'].font = Font(name='Cambria', size=9, italic=True, color='718096')

# Row 3: Tick counter
ws['A3'].value = 'TICK'
ws['A3'].font = SF
ws['B3'].value = '=B3+1'  # circular
ws['B3'].font = Font(name='Consolas', size=14, bold=True, color='9B2C2C')
ws['C3'].value = 'Active(>0.5):'
ws['C3'].font = BF
act_range = f'{ncol(0)}7:{ncol(NN-1)}7'
ws['F3'].value = f'=COUNTIF({act_range},">"&0.5)'
ws['F3'].font = FF
ws['G3'].value = f'/{NN}'
ws['G3'].font = BF
ws['H3'].value = 'Avg:'
ws['H3'].font = BF
ws['I3'].value = f'=ROUND(AVERAGE({act_range}),4)'
ws['I3'].font = FF

# Row 4: blank separator
# Row 5: INPUT — plain cells, yellow, editable, NO MERGE
ws['A5'].value = 'INPUT'
ws['A5'].font = Font(name='Cambria', size=11, bold=True, color='9B2C2C')

# Pre-fill alternating pattern so it starts with activity
for j in range(NI):
    c = ws.cell(row=5, column=j+2)
    c.value = 1.0 if j % 2 == 0 else 0.0  # alternating pattern pre-loaded
    c.font = Font(name='Consolas', size=11, bold=True, color='9B2C2C')
    c.alignment = C
    c.fill = INPUT_FILL
    c.border = TB

# Fill non-input columns in row 5 with grey
for j in range(NI, NN):
    c = ws.cell(row=5, column=j+2)
    c.value = ''
    c.fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

# Row 6: blank
# Row 7: ACTIVATION — THE LIVE ROW with circular refs
ws['A7'].value = 'ACTIVATION'
ws['A7'].font = SF

# Generate polarity
polarities = []
for j in range(NN):
    if j < NI or j >= NI+NH:
        polarities.append(1.0)
    else:
        polarities.append(-1.0 if random.random() < 0.2 else 1.0)

for j in range(NN):
    c = ws.cell(row=7, column=j+2)
    c.alignment = C; c.border = TB; c.number_format = '0.00'
    if j < NI:
        c.value = f'={ncol(j)}5'  # clamp to input
        c.font = FF; c.fill = INPUT_FILL
    else:
        w_range = f'{ncol(j)}$15:{ncol(j)}${14+NN}'
        a_range = f'${ncol(0)}$7:${ncol(NN-1)}$7'
        p_range = f'${ncol(0)}$9:${ncol(NN-1)}$9'
        g_cell = f'{ncol(j)}8'
        c.value = f'=1/(1+EXP(-{g_cell}*SUMPRODUCT({w_range},{a_range},{p_range})))'
        c.font = Font(name='Consolas', size=8, bold=True, color='1A365D')
        if j >= NI+NH:
            c.fill = OUT_FILL
        else:
            c.fill = ACT_FILL

# Row 8: GAIN
ws['A8'].value = 'GAIN'
ws['A8'].font = SF
for j in range(NN):
    c = ws.cell(row=8, column=j+2)
    c.alignment = C; c.border = TB; c.number_format = '0.00'
    if j < NI:
        c.value = 1.0; c.font = BF
    else:
        avg_cell = f'{ncol(j)}10'
        c.value = f'=MAX(0.1,MIN(5,{ncol(j)}8+0.02*(0.3-{avg_cell})))'
        c.font = FF

# Row 9: POLARITY
ws['A9'].value = 'POLARITY'
ws['A9'].font = SF
for j in range(NN):
    p = polarities[j]
    c = ws.cell(row=9, column=j+2, value=p)
    c.font = BF; c.alignment = C; c.border = TB
    c.fill = EXC_FILL if p > 0 else INH_FILL

# Row 10: ACTIVITY AVG (circular EMA)
ws['A10'].value = 'ACT AVG'
ws['A10'].font = SF
for j in range(NN):
    c = ws.cell(row=10, column=j+2)
    c.alignment = C; c.border = TB; c.number_format = '0.000'
    act_cell = f'{ncol(j)}7'
    c.value = f'=0.9*{ncol(j)}10+0.1*{act_cell}'  # circular EMA
    c.font = FF

# Row 11: ROLE labels
ws['A11'].value = 'NEURON'
ws['A11'].font = SF
for j in range(NN):
    c = ws.cell(row=11, column=j+2, value=nlabel(j))
    c.font = HF; c.fill = HFILL; c.alignment = C; c.border = TB

# Row 12: blank
# Row 13: OUTPUT readout (just the output neurons, bigger font)
ws['A13'].value = 'OUTPUT'
ws['A13'].font = Font(name='Cambria', size=12, bold=True, color='553C9A')
for j in range(NO):
    out_j = NI + NH + j
    c = ws.cell(row=13, column=j+2)
    c.value = f'={ncol(out_j)}7'
    c.font = Font(name='Consolas', size=12, bold=True, color='553C9A')
    c.alignment = C; c.number_format = '0.000'
    c.fill = OUT_FILL; c.border = TB

ws.cell(row=14, column=1, value='(O0-O19)').font = Font(name='Cambria', size=8, italic=True, color='718096')

# ═══ WEIGHT MATRIX starts at row 15 ═══
ws['A15'].value = 'W[i→j]'
ws['A15'].font = SF

# Column headers row 15
# Actually put headers in row 15, data in rows 16-115
# Wait — the formulas reference rows 15 to 15+NN-1 for weights
# Let me put the weight data directly at row 15 to match the formula references

# Row headers + weights
for i in range(NN):
    r = 15 + i
    c = ws.cell(row=r, column=1, value=nlabel(i))
    c.font = HF; c.fill = HFILL; c.alignment = C; c.border = TB

    for j in range(NN):
        c = ws.cell(row=r, column=j+2)
        c.alignment = C; c.border = TB; c.number_format = '0.00'
        c.font = Font(name='Consolas', size=7, color='4A5568')

        weight = 0.0
        if i < NI and NI <= j < NI+NH:
            if random.random() < 0.30:
                weight = round(random.gauss(0.3, 0.5), 2)
        elif NI <= i < NI+NH and NI <= j < NI+NH and i != j:
            if random.random() < 0.10:
                weight = round(random.gauss(0.0, 0.6), 2)
        elif NI <= i < NI+NH and j >= NI+NH:
            if random.random() < 0.30:
                weight = round(random.gauss(0.3, 0.5), 2)
        elif j >= NI+NH and NI <= i < NI+NH:  # output→hidden feedback
            if random.random() < 0.03:
                weight = round(random.gauss(0.0, 0.3), 2)

        c.value = weight
        if weight > 0.01: c.fill = EXC_FILL
        elif weight < -0.01: c.fill = INH_FILL

# ═══ ACTIVATION BAR CHART below weight matrix ═══
chart_row = 15 + NN + 2
chart = BarChart()
chart.title = "Live Activation (press F9)"
chart.y_axis.title = "Activation"
chart.y_axis.scaling.min = 0; chart.y_axis.scaling.max = 1
chart.style = 13; chart.width = 40; chart.height = 12

cats = Reference(ws, min_col=2, max_col=NN+1, min_row=11)
vals = Reference(ws, min_col=2, max_col=NN+1, min_row=7)
chart.add_data(vals, from_rows=True, titles_from_data=False)
chart.set_categories(cats)

ws.add_chart(chart, f'A{chart_row}')

# ═══ TEST PATTERNS ═══
pat_row = chart_row + 18
ws.cell(row=pat_row, column=1, value='TEST PATTERNS').font = SF
ws.cell(row=pat_row+1, column=1, value='Copy a row into INPUT (row 5), then press F9 20x').font = BF

patterns = [
    ('Alternating A', [1,0]*8),
    ('Alternating B', [0,1]*8),
    ('Left half on',  [1]*8+[0]*8),
    ('Right half on', [0]*8+[1]*8),
    ('All on',        [1]*16),
    ('All off',       [0]*16),
    ('Binary 170',    [1,0,1,0,1,0,1,0,1,0,1,0,1,0,1,0]),
    ('Center on',     [0,0,0,0,1,1,1,1,1,1,1,1,0,0,0,0]),
    ('Edges only',    [1,1,0,0,0,0,0,0,0,0,0,0,0,0,1,1]),
    ('Random',        [random.randint(0,1) for _ in range(16)]),
]
for idx, (name, pat) in enumerate(patterns):
    r = pat_row + 2 + idx
    ws.cell(row=r, column=1, value=name).font = BF
    for j, v in enumerate(pat):
        ws.cell(row=r, column=j+2, value=v).font = FF
        ws.cell(row=r, column=j+2).alignment = C

# ═══ STATS ═══
stat_row = pat_row + 2 + len(patterns) + 2
ws.cell(row=stat_row, column=1, value='NETWORK STATS').font = SF
stats = [
    ('Neurons:', NN),
    ('Excitatory:', f'=COUNTIF({ncol(0)}9:{ncol(NN-1)}9,1)'),
    ('Inhibitory:', f'=COUNTIF({ncol(0)}9:{ncol(NN-1)}9,-1)'),
    ('Connections:', f'=COUNTIF({ncol(0)}15:{ncol(NN-1)}{14+NN},"<>0")'),
    ('Avg gain:', f'=ROUND(AVERAGE({ncol(0)}8:{ncol(NN-1)}8),3)'),
]
for i, (label, val) in enumerate(stats):
    ws.cell(row=stat_row+1+i, column=1, value=label).font = BF
    c = ws.cell(row=stat_row+1+i, column=2)
    c.value = val; c.font = FF

# ═══ INSTRUCTIONS ═══
inst_row = stat_row + len(stats) + 3
insts = [
    'HOW IT WORKS:',
    '1. Row 5 (yellow) = INPUT. Edit these values (0 or 1).',
    '2. Row 7 = ACTIVATION. Each hidden/output neuron computes:',
    '   sigmoid(gain * sum(weights * activations * polarities))',
    '3. The formulas are CIRCULAR — activation references itself.',
    '4. Enable: File > Options > Formulas > Enable iterative calculation',
    '5. Set Max Iterations = 1 for step-by-step, or 100 for fast settling.',
    '6. Press F9. Each press = one network tick.',
    '7. Row 8 (GAIN) adapts homeostatically: overactive neurons reduce gain.',
    '8. Row 10 (ACT AVG) tracks exponential moving average of activity.',
    '9. Row 9 (POLARITY): green=excitatory(+1), red=inhibitory(-1).',
    '10. Row 13 = OUTPUT readout (purple). Different inputs → different outputs.',
]
for i, line in enumerate(insts):
    ws.cell(row=inst_row+i, column=1, value=line).font = SF if i == 0 else FF

print(f"  {NN} neurons, {NN}x{NN} matrix, pre-loaded alternating input")

# Save
out = r'D:\claude\proofs\Deficiency_Research_Instrument.xlsx'
wb.save(out)
print(f"Saved: {out}")
