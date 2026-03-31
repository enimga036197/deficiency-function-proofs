"""
Rebuild the Neural Substrate sheet as a LIVE recurrent network.
No VBA. Circular references + iterative calculation = each F9 is one tick.

Architecture: 50 neurons (10 input, 30 hidden, 10 output)
- Weight matrix: 50x50 sparse connections
- Activation row: CIRCULAR REFERENCES — each neuron reads all activations
- Gain row: homeostatic — adapts toward target activity
- Activity average: exponential moving average (circular ref)
- E/I polarity: ~20% inhibitory
- Input clamping: input neurons read from editable input cells
- Tick counter: increments each recalc (circular ref)
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
import random, math

random.seed(77)

# Styles
TF = Font(name='Cambria', size=14, bold=True, color='1B2A4A')
SF = Font(name='Cambria', size=11, bold=True, color='2D3748')
HF = Font(name='Cambria', size=8, bold=True, color='FFFFFF')
BF = Font(name='Cambria', size=8, color='2D3748')
FF = Font(name='Consolas', size=8, color='1A365D')
RF = Font(name='Cambria', size=9, bold=True, color='276749')
INPUT_F = Font(name='Consolas', size=10, bold=True, color='9B2C2C')
HFILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
EXC_FILL = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')
INH_FILL = PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FEFCBF', end_color='FEFCBF', fill_type='solid')
ACT_FILL = PatternFill(start_color='EBF8FF', end_color='EBF8FF', fill_type='solid')
TB = Border(left=Side('thin','E2E8F0'), right=Side('thin','E2E8F0'),
            top=Side('thin','E2E8F0'), bottom=Side('thin','E2E8F0'))
C = Alignment(horizontal='center', vertical='center')

NN = 50   # Total neurons
NI = 10   # Input
NH = 30   # Hidden
NO = 10   # Output

print(f"Building live substrate: {NN} neurons ({NI}in, {NH}hid, {NO}out)...")

wb = openpyxl.load_workbook(r'D:\claude\proofs\Deficiency_Research_Instrument.xlsx')

# Delete old substrate sheet if exists
if 'Neural Substrate' in wb.sheetnames:
    del wb['Neural Substrate']

ws = wb.create_sheet('Neural Substrate', 8)  # Insert at position 9
ws.sheet_properties.tabColor = '9B2C2C'

# Column widths: col A=labels (wider), rest=narrow for matrix
ws.column_dimensions['A'].width = 12
for j in range(NN):
    ws.column_dimensions[get_column_letter(j+2)].width = 4.5

# ── ROW LAYOUT ──
# Row 1: Title
# Row 2: Subtitle
# Row 3: Tick counter + instructions
# Row 4: blank
# Row 5: INPUT PATTERN (editable, yellow) — only NI cells
# Row 6: blank
# Row 7: ACTIVATION (circular refs — THE LIVE ROW)
# Row 8: GAIN (circular refs — homeostatic adaptation)
# Row 9: ACTIVITY AVG (circular refs — EMA)
# Row 10: POLARITY (+1 or -1)
# Row 11: ROLE LABEL (IN/HID/OUT)
# Row 12: blank
# Row 13: WEIGHT MATRIX header
# Row 14: Column headers
# Row 15-64: Weight matrix (50 rows)
# Row 66+: Output readout, patterns, charts

R_TITLE = 1
R_SUB = 2
R_TICK = 3
R_INPUT = 5
R_ACT = 7
R_GAIN = 8
R_AVG = 9
R_POL = 10
R_ROLE = 11
R_WHDR = 13
R_CHDR = 14
R_WSTART = 15
R_WEND = R_WSTART + NN - 1

# Helper: neuron label
def nlabel(j):
    if j < NI: return f'I{j}'
    elif j < NI+NH: return f'H{j-NI}'
    else: return f'O{j-NI-NH}'

# Helper: column letter for neuron j (0-indexed)
def ncol(j): return get_column_letter(j + 2)

# ── TITLE ──
ws.merge_cells(f'A{R_TITLE}:{ncol(NN-1)}{R_TITLE}')
ws[f'A{R_TITLE}'].value = f'NEURAL SUBSTRATE — {NN}-neuron live recurrent network | Each F9 = one tick'
ws[f'A{R_TITLE}'].font = TF; ws[f'A{R_TITLE}'].alignment = C

ws.merge_cells(f'A{R_SUB}:{ncol(NN-1)}{R_SUB}')
ws[f'A{R_SUB}'].value = f'{NI} input → {NH} hidden → {NO} output | Sigmoid | Circular refs = live dynamics | Enable: File > Options > Formulas > Iterative Calculation'
ws[f'A{R_SUB}'].font = Font(name='Cambria', size=9, italic=True, color='4A5568')
ws[f'A{R_SUB}'].alignment = C

# ── TICK COUNTER (circular ref: increments by 1 each recalc) ──
ws[f'A{R_TICK}'].value = 'Tick:'
ws[f'A{R_TICK}'].font = SF
ws[f'B{R_TICK}'].value = f'=B{R_TICK}+1'  # CIRCULAR — increments each F9
ws[f'B{R_TICK}'].font = Font(name='Consolas', size=14, bold=True, color='9B2C2C')
ws[f'C{R_TICK}'].value = 'Active neurons:'
ws[f'C{R_TICK}'].font = BF
# Count neurons with activation > 0.5
act_range = f'{ncol(0)}{R_ACT}:{ncol(NN-1)}{R_ACT}'
ws[f'E{R_TICK}'].value = f'=COUNTIF({act_range},">"&0.5)&"/{NN}"'
ws[f'E{R_TICK}'].font = FF
ws[f'F{R_TICK}'].value = 'Avg activation:'
ws[f'F{R_TICK}'].font = BF
ws[f'H{R_TICK}'].value = f'=ROUND(AVERAGE({act_range}),3)'
ws[f'H{R_TICK}'].font = FF

# ── INPUT PATTERN ROW ──
ws[f'A{R_INPUT}'].value = 'INPUT →'
ws[f'A{R_INPUT}'].font = Font(name='Cambria', size=10, bold=True, color='9B2C2C')
for j in range(NI):
    c = ws.cell(row=R_INPUT, column=j+2, value=0.0)
    c.font = INPUT_F; c.alignment = C; c.fill = INPUT_FILL; c.border = TB
# Label the non-input columns
for j in range(NI, NN):
    c = ws.cell(row=R_INPUT, column=j+2, value='')
    c.fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

# ── ACTIVATION ROW (THE LIVE PART) ──
ws[f'A{R_ACT}'].value = 'ACTIVATION'
ws[f'A{R_ACT}'].font = SF

# Generate polarity: ~20% inhibitory for hidden neurons
polarities = []
for j in range(NN):
    if j < NI:
        polarities.append(1.0)  # inputs always excitatory
    elif j >= NI + NH:
        polarities.append(1.0)  # outputs always excitatory
    else:
        polarities.append(-1.0 if random.random() < 0.2 else 1.0)

# Generate initial gains
gains = [1.0] * NN

for j in range(NN):
    col = j + 2
    c = ws.cell(row=R_ACT, column=col)
    c.alignment = C; c.border = TB; c.number_format = '0.00'

    if j < NI:
        # Input neurons: clamp to input row
        c.value = f'={ncol(j)}{R_INPUT}'
        c.font = FF; c.fill = INPUT_FILL
    else:
        # Hidden/Output: sigmoid(gain * sum(W * activation * polarity))
        # W column j is in rows R_WSTART to R_WEND, column j+2
        # Activation row is R_ACT, columns 2 to NN+1
        w_range = f'{ncol(j)}${R_WSTART}:{ncol(j)}${R_WEND}'
        a_range = f'${ncol(0)}${R_ACT}:${ncol(NN-1)}${R_ACT}'
        p_range = f'${ncol(0)}${R_POL}:${ncol(NN-1)}${R_POL}'
        g_cell = f'{ncol(j)}{R_GAIN}'

        # The formula: =1/(1+EXP(-gain * SUMPRODUCT(weights, activations, polarities)))
        # This is CIRCULAR because a_range includes this cell
        c.value = f'=1/(1+EXP(-{g_cell}*SUMPRODUCT({w_range},{a_range},{p_range})))'
        c.font = Font(name='Consolas', size=8, bold=True, color='1A365D')
        c.fill = ACT_FILL

# ── GAIN ROW (homeostatic — circular ref) ──
ws[f'A{R_GAIN}'].value = 'GAIN'
ws[f'A{R_GAIN}'].font = SF
TARGET_ACT = 0.3
HOMEO_RATE = 0.02

for j in range(NN):
    col = j + 2
    c = ws.cell(row=R_GAIN, column=col)
    c.alignment = C; c.border = TB; c.number_format = '0.00'

    if j < NI:
        c.value = 1.0; c.font = BF
    else:
        # Homeostatic: gain += rate * (target - activity_avg)
        # Clamped to [0.1, 5.0]
        avg_cell = f'{ncol(j)}{R_AVG}'
        # CIRCULAR: references itself
        c.value = f'=MAX(0.1,MIN(5,{ncol(j)}{R_GAIN}+{HOMEO_RATE}*({TARGET_ACT}-{avg_cell})))'
        c.font = FF

# ── ACTIVITY AVERAGE ROW (EMA — circular ref) ──
ws[f'A{R_AVG}'].value = 'ACT AVG'
ws[f'A{R_AVG}'].font = SF
DECAY = 0.9

for j in range(NN):
    col = j + 2
    c = ws.cell(row=R_AVG, column=col)
    c.alignment = C; c.border = TB; c.number_format = '0.000'
    act_cell = f'{ncol(j)}{R_ACT}'
    # EMA: avg = decay * avg + (1-decay) * current — CIRCULAR
    c.value = f'={DECAY}*{ncol(j)}{R_AVG}+{1-DECAY}*{act_cell}'
    c.font = FF

# ── POLARITY ROW ──
ws[f'A{R_POL}'].value = 'POLARITY'
ws[f'A{R_POL}'].font = SF
for j in range(NN):
    col = j + 2
    p = polarities[j]
    c = ws.cell(row=R_POL, column=col, value=p)
    c.font = BF; c.alignment = C; c.border = TB
    c.fill = EXC_FILL if p > 0 else INH_FILL

# ── ROLE ROW ──
ws[f'A{R_ROLE}'].value = 'ROLE'
ws[f'A{R_ROLE}'].font = SF
for j in range(NN):
    label = nlabel(j)
    c = ws.cell(row=R_ROLE, column=j+2, value=label)
    c.font = HF; c.fill = HFILL; c.alignment = C; c.border = TB

# ── WEIGHT MATRIX ──
ws[f'A{R_WHDR}'].value = 'WEIGHT MATRIX W[i→j]'
ws[f'A{R_WHDR}'].font = SF

# Column headers
for j in range(NN):
    c = ws.cell(row=R_CHDR, column=j+2, value=nlabel(j))
    c.font = HF; c.fill = HFILL; c.alignment = C; c.border = TB

# Row headers + weight values
for i in range(NN):
    r = R_WSTART + i
    c = ws.cell(row=r, column=1, value=nlabel(i))
    c.font = HF; c.fill = HFILL; c.alignment = C; c.border = TB

    for j in range(NN):
        c = ws.cell(row=r, column=j+2)
        c.alignment = C; c.border = TB; c.number_format = '0.00'
        c.font = Font(name='Consolas', size=7, color='2D3748')

        weight = 0.0

        # Input → Hidden: ~35% connected, moderate weights
        if i < NI and NI <= j < NI+NH:
            if random.random() < 0.35:
                weight = round(random.gauss(0.3, 0.4), 2)

        # Hidden → Hidden: ~15% connected, lateral (both E and I)
        elif NI <= i < NI+NH and NI <= j < NI+NH and i != j:
            if random.random() < 0.15:
                weight = round(random.gauss(0.0, 0.5), 2)

        # Hidden → Output: ~35% connected
        elif NI <= i < NI+NH and j >= NI+NH:
            if random.random() < 0.35:
                weight = round(random.gauss(0.3, 0.4), 2)

        # Output → Hidden (feedback): ~5% connected, weak
        elif j >= NI+NH and NI <= i < NI+NH:
            if random.random() < 0.05:
                weight = round(random.gauss(0.0, 0.2), 2)

        c.value = weight
        if weight > 0.01:
            c.fill = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')
        elif weight < -0.01:
            c.fill = PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid')

# ── OUTPUT READOUT ──
R_OUT = R_WEND + 2
ws[f'A{R_OUT}'].value = 'OUTPUT'
ws[f'A{R_OUT}'].font = Font(name='Cambria', size=12, bold=True, color='276749')
for j in range(NO):
    out_j = NI + NH + j
    c = ws.cell(row=R_OUT, column=j+2)
    c.value = f'={ncol(out_j)}{R_ACT}'
    c.font = Font(name='Consolas', size=12, bold=True, color='276749')
    c.alignment = C; c.number_format = '0.000'
    c.fill = ACT_FILL

# Output labels
for j in range(NO):
    ws.cell(row=R_OUT+1, column=j+2, value=f'O{j}').font = HF
    ws.cell(row=R_OUT+1, column=j+2).fill = HFILL
    ws.cell(row=R_OUT+1, column=j+2).alignment = C

# ── ACTIVATION BAR CHART ──
chart = BarChart()
chart.title = "Neuron Activation Levels (live)"
chart.y_axis.title = "Activation"
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 1
chart.style = 13; chart.width = 30; chart.height = 10

cats = Reference(ws, min_col=2, max_col=NN+1, min_row=R_ROLE)
vals = Reference(ws, min_col=2, max_col=NN+1, min_row=R_ACT)
chart.add_data(vals, from_rows=True, titles_from_data=False)
chart.set_categories(cats)
chart.shape = 4

ws.add_chart(chart, f'A{R_OUT+3}')

# ── TEST PATTERNS ──
R_PAT = R_OUT + 20
ws[f'A{R_PAT}'].value = 'TEST PATTERNS (copy into yellow INPUT cells, then press F9 repeatedly)'
ws[f'A{R_PAT}'].font = SF

patterns = [
    ('Alternating A', [1,0,1,0,1,0,1,0,1,0]),
    ('Alternating B', [0,1,0,1,0,1,0,1,0,1]),
    ('Left half',     [1,1,1,1,1,0,0,0,0,0]),
    ('Right half',    [0,0,0,0,0,1,1,1,1,1]),
    ('Edges',         [1,0,0,0,0,0,0,0,0,1]),
    ('Center',        [0,0,0,1,1,1,1,0,0,0]),
    ('All on',        [1,1,1,1,1,1,1,1,1,1]),
    ('All off',       [0,0,0,0,0,0,0,0,0,0]),
    ('Binary 42',     [0,0,1,0,1,0,1,0,0,0]),
    ('Binary 170',    [1,0,1,0,1,0,1,0,1,0]),
]

for idx, (name, pat) in enumerate(patterns):
    r = R_PAT + 1 + idx
    ws.cell(row=r, column=1, value=name).font = BF
    for j, v in enumerate(pat):
        c = ws.cell(row=r, column=j+2, value=v)
        c.font = FF; c.alignment = C

# ── INSTRUCTIONS ──
R_INST = R_PAT + len(patterns) + 2
instructions = [
    'HOW TO USE THIS LIVE NETWORK:',
    '',
    '1. ENABLE ITERATIVE CALCULATION:',
    '   File → Options → Formulas → check "Enable iterative calculation"',
    '   Set Maximum Iterations = 1, Maximum Change = 0.001',
    '   This makes circular references iterate ONCE per F9 press = one network tick.',
    '',
    '2. SET INPUT: Copy a test pattern into the yellow INPUT cells (row 5).',
    '',
    '3. PRESS F9: Each press propagates signals one step through the network.',
    '   Watch the ACTIVATION row and OUTPUT row change.',
    '   Watch the bar chart update in real time.',
    '   The tick counter increments each press.',
    '',
    '4. OBSERVE: Different input patterns produce different output signatures.',
    '   The network settles into attractors — press F9 20+ times to reach equilibrium.',
    '   Homeostatic gain adapts over time: overactive neurons decrease gain, quiet ones increase.',
    '',
    '5. THE NETWORK IS ALIVE: Every formula references other formulas.',
    '   Activation depends on weights and other activations (circular).',
    '   Gain depends on activity average (circular). Activity average depends on activation (circular).',
    '   The tick counter counts itself (circular). No VBA. Pure Excel dynamics.',
]
for i, line in enumerate(instructions):
    ws.cell(row=R_INST+i, column=1, value=line).font = BF if line and not line.startswith(' ') else FF

# ── NETWORK STATISTICS ──
R_STAT = R_INST + len(instructions) + 1
ws[f'A{R_STAT}'].value = 'NETWORK STATISTICS (live)'; ws[f'A{R_STAT}'].font = SF

stats = [
    ('Total neurons:', f'={NN}'),
    ('Excitatory:', f'=COUNTIF({ncol(0)}{R_POL}:{ncol(NN-1)}{R_POL},1)'),
    ('Inhibitory:', f'=COUNTIF({ncol(0)}{R_POL}:{ncol(NN-1)}{R_POL},-1)'),
    ('E/I ratio:', f'=ROUND(COUNTIF({ncol(0)}{R_POL}:{ncol(NN-1)}{R_POL},1)/COUNTIF({ncol(0)}{R_POL}:{ncol(NN-1)}{R_POL},-1),1)&":1"'),
    ('Connections:', f'=COUNTIF({ncol(0)}{R_WSTART}:{ncol(NN-1)}{R_WEND},"<>"&0)'),
    ('Avg gain:', f'=ROUND(AVERAGE({ncol(0)}{R_GAIN}:{ncol(NN-1)}{R_GAIN}),3)'),
    ('Avg activity:', f'=ROUND(AVERAGE({ncol(0)}{R_AVG}:{ncol(NN-1)}{R_AVG}),4)'),
    ('Max activation:', f'=ROUND(MAX({ncol(0)}{R_ACT}:{ncol(NN-1)}{R_ACT}),3)'),
    ('Min activation:', f'=ROUND(MIN({ncol(NI)}{R_ACT}:{ncol(NN-1)}{R_ACT}),3)'),
]
for i, (label, formula) in enumerate(stats):
    ws.cell(row=R_STAT+1+i, column=1, value=label).font = BF
    ws.cell(row=R_STAT+1+i, column=2).value = formula
    ws.cell(row=R_STAT+1+i, column=2).font = FF

print(f"  Built: {NN} neurons, {NN}x{NN} weight matrix")
print(f"  Circular refs: activation, gain, activity avg, tick counter")
print(f"  {len(patterns)} test patterns")

# Save
out = r'D:\claude\proofs\Deficiency_Research_Instrument.xlsx'
wb.save(out)
print(f"\nSaved: {out}")
print("Open in Excel, enable iterative calculation, and press F9!")
